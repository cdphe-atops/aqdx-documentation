import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from ruamel.yaml import YAML


def extract_comment(node, key):
    """Safely extracts the inline comment for a given key, ignoring block comments."""
    try:
        comments = node.ca.items.get(key)
        # Inline comments are typically in the 3rd position of the comment array
        if comments and len(comments) >= 3 and comments[2]:
            # Split by '\n' and take [0] to strictly grab the inline text,
            # leaving behind any "# =====" section headers grouped by ruamel.
            raw_comment = comments[2].value.split("\n")[0].strip("# \n")
            return raw_comment
    except (AttributeError, TypeError):
        pass
    return ""


def split_instruction(comment):
    """Splits the YAML comment into Requirement, Format, and Description."""
    parts = [p.strip() for p in comment.split("|")]
    if not parts or parts == [""]:
        return "", "", ""

    req_rules = []
    format_rule = ""
    desc_rule = ""

    for part in parts:
        part_lower = part.lower()
        if part_lower.startswith("format:"):
            format_rule = part
        elif part_lower.startswith("desc:"):
            desc_rule = part[5:].strip()  # Removes the "Desc:" tag
        else:
            req_rules.append(part)

    return " | ".join(req_rules), format_rule, desc_rule


def format_instruction_rows(ws, num_rows=3):
    """Applies gray background, italics, and freezes panes for horizontal tabs."""
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    italic_font = Font(italic=True, color="555555")

    # Format instruction rows (1 to num_rows)
    for row in range(1, num_rows + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = gray_fill
            cell.font = italic_font

    # Bold the actual headers (the row immediately following instructions)
    header_row = num_rows + 1
    for col in range(1, ws.max_column + 1):
        ws.cell(row=header_row, column=col).font = Font(bold=True)

    # Freeze panes so user scrolling doesn't hide instructions or headers
    # If headers are row 4, freeze at row 5
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def autofit_columns(ws):
    """Iterates through all columns in a worksheet and adjusts the width."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Add a little extra padding (e.g., +2) so it doesn't look cramped
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


def main():
    yaml = YAML()
    yaml.preserve_quotes = True

    # Load your YAML template
    try:
        with open("../../docs/assets/AQDx_metadata_form_v3.yaml", "r") as f:
            data = yaml.load(f)
    except FileNotFoundError:
        print(
            "Error: Could not find 'AQDx_metadata_form_v3.yaml'. Please ensure it is in the same directory."
        )
        sys.exit(1)

    wb = Workbook()

    # --------------------------------------------------------------------------
    # TAB 1: Project_Info (Vertical Layout)
    # --------------------------------------------------------------------------
    ws_proj = wb.active
    ws_proj.title = "Project_Info"
    # Expanded headers for Tab 1 to match the new 3-part split
    ws_proj.append(
        ["Metadata Field", "Response", "Requirement", "Format", "Description"]
    )

    proj_sections = [
        (data, ["dataset_id", "aqdx_metadata_version", "aqdx_data_version"]),
        (data["data_steward"], data["data_steward"].keys()),
        (data["dataset_quality"], data["dataset_quality"].keys()),
    ]

    for node, keys in proj_sections:
        for key in keys:
            comment = extract_comment(node, key)
            req, fmt, desc = split_instruction(comment)
            ws_proj.append([key, "", req, fmt, desc])

    for col in range(1, 6):
        ws_proj.cell(row=1, column=col).font = Font(bold=True)
    ws_proj.freeze_panes = "A2"

    # --------------------------------------------------------------------------
    # TAB 2: Sites (Horizontal Layout)
    # --------------------------------------------------------------------------
    ws_sites = wb.create_sheet(title="Sites")
    site_node = data["sites"][0]

    reqs, fmts, descs, headers = [], [], [], []
    for key in site_node.keys():
        comment = extract_comment(site_node, key)
        req, fmt, desc = split_instruction(comment)
        reqs.append(req)
        fmts.append(fmt)
        descs.append(desc)
        headers.append(key)

    ws_sites.append(reqs)
    ws_sites.append(fmts)
    ws_sites.append(descs)
    ws_sites.append(headers)

    format_instruction_rows(ws_sites, num_rows=3)

    # --------------------------------------------------------------------------
    # TAB 3: Instruments (Horizontal Layout)
    # --------------------------------------------------------------------------
    ws_inst = wb.create_sheet(title="Instruments")
    inst_node = data["instruments"][0]

    reqs, fmts, descs, headers = [], [], [], []
    for key in inst_node.keys():
        if key == "parameters":
            continue
        comment = extract_comment(inst_node, key)
        req, fmt, desc = split_instruction(comment)
        reqs.append(req)
        fmts.append(fmt)
        descs.append(desc)
        headers.append(key)

    ws_inst.append(reqs)
    ws_inst.append(fmts)
    ws_inst.append(descs)
    ws_inst.append(headers)

    format_instruction_rows(ws_inst, num_rows=3)

    # Data Validation for site_name (Column B in Instruments -> Sites Tab)
    # Headers are Row 4, Data starts Row 5
    dv_site = DataValidation(
        type="list", formula1="=Sites!$A$5:$A$1000", allow_blank=True
    )
    ws_inst.add_data_validation(dv_site)
    dv_site.add("B5:B1000")

    # --------------------------------------------------------------------------
    # TAB 4: Parameters (Horizontal Layout)
    # --------------------------------------------------------------------------
    ws_params = wb.create_sheet(title="Parameters")
    param_node = data["instruments"][0]["parameters"][0]

    reqs = ["REQUIRED | Matches tabular data"]
    fmts = ["Format: string"]
    descs = [""]
    headers = ["device_id"]

    for key in param_node.keys():
        comment = extract_comment(param_node, key)
        req, fmt, desc = split_instruction(comment)
        reqs.append(req)
        fmts.append(fmt)
        descs.append(desc)
        headers.append(key)

    ws_params.append(reqs)
    ws_params.append(fmts)
    ws_params.append(descs)
    ws_params.append(headers)

    format_instruction_rows(ws_params, num_rows=3)

    # Data Validation for device_id (Column A in Parameters -> Instruments Tab)
    dv_device = DataValidation(
        type="list", formula1="=Instruments!$A$5:$A$1000", allow_blank=True
    )
    ws_params.add_data_validation(dv_device)
    dv_device.add("A5:A1000")

    # --------------------------------------------------------------------------
    # Auto-fit columns for all sheets
    # --------------------------------------------------------------------------
    autofit_columns(ws_proj)
    autofit_columns(ws_sites)
    autofit_columns(ws_inst)
    autofit_columns(ws_params)

    # --------------------------------------------------------------------------
    # Save the Workbook
    # --------------------------------------------------------------------------
    output_filename = "AQDx_metadata_template.xlsx"
    wb.save(output_filename)
    print(f"Spreadsheet generated successfully: {output_filename}")


if __name__ == "__main__":
    main()
